import base64
import csv
import io
from datetime import datetime
from itertools import groupby
from textwrap import wrap
from typing import Iterable

import matplotlib
import numpy as np
import pdfkit
from jinja2 import Environment, FileSystemLoader
from matplotlib import pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border
from openpyxl.worksheet.worksheet import Worksheet

matplotlib.use('TkAgg')

WKHTMLTOPDF_PATH = r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'


class Salary:
    def __init__(self, from_amount, to_amount, currency):
        self.from_amount = from_amount
        self.to_amount = to_amount
        self.currency = currency

    @property
    def avg_ruble_amount(self):
        currency_to_rub = {
            "AZN": 35.68,
            "BYR": 23.91,
            "EUR": 59.90,
            "GEL": 21.74,
            "KGS": 0.76,
            "KZT": 0.13,
            "RUR": 1,
            "UAH": 1.64,
            "USD": 60.66,
            "UZS": 0.0055,
        }

        avg = (self.from_amount + self.to_amount) / 2
        return avg * currency_to_rub[self.currency]

    @staticmethod
    def parse_from_dict(row_dict):
        return Salary(
            from_amount=int(float(row_dict['salary_from'])),
            to_amount=int(float(row_dict['salary_to'])),
            currency=row_dict['salary_currency']
        )


class Vacancy:
    def __init__(self, name, salary: Salary, area_name, published_at: datetime):
        self.name = name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at

    @staticmethod
    def parse_from_dict(row_dict):
        return Vacancy(
            name=row_dict['name'],
            salary=Salary.parse_from_dict(row_dict),
            area_name=row_dict['area_name'],
            published_at=datetime.strptime(row_dict['published_at'], '%Y-%m-%dT%H:%M:%S%z')
        )

    @property
    def year(self):
        return self.published_at.year


class AvgCounter:
    def __init__(self, first_value=None):
        self.value = 0
        self._sum = 0
        self._len = 0
        if first_value is not None:
            self.add(first_value)

    def add(self, value):
        self._sum += value
        self._len += 1
        self.value = self._sum / self._len


class VacanciesStatistics:
    def __init__(self, file_name, prof_name):
        self.file_name = file_name
        self.prof_name = prof_name

        self._salaries_by_year = {}
        self.counts_by_year = {}
        self._prof_salaries_by_year = {}
        self._prof_counts_by_year = {}
        self._salaries_by_cities = {}

        self._counts_by_cities = {}
        self._vacancies_count = 0
        self._init_data()

    def _init_data(self):
        with open(self.file_name, 'r', encoding='utf_8_sig') as file:
            reader = csv.reader(file)
            title_row = None
            for row in reader:
                if title_row is None:
                    title_row = row
                    continue
                row_dict = self._to_row_dict(title_row, row)
                if row_dict is None:
                    continue
                vacancy = Vacancy.parse_from_dict(row_dict)
                self._update_statistics(vacancy)

    def _update_statistics(self, vacancy):
        def change_salary_and_count_stats(salaries: dict, counts: dict, key, salary):
            if key not in salaries:
                salaries[key] = AvgCounter(salary)
                counts[key] = 1
            else:
                salaries[key].add(salary)
                counts[key] += 1

        salary = vacancy.salary.avg_ruble_amount
        year = vacancy.year
        # Year stats
        change_salary_and_count_stats(self._salaries_by_year, self.counts_by_year, year, salary)

        # City stats
        self._vacancies_count += 1
        city = vacancy.area_name
        change_salary_and_count_stats(self._salaries_by_cities, self._counts_by_cities, city, salary)

        # Prof stats
        if self.prof_name not in vacancy.name:
            return
        change_salary_and_count_stats(self._prof_salaries_by_year, self._prof_counts_by_year, year, salary)

    @property
    def salaries_by_year(self):
        return dict(map(lambda item: (item[0], int(item[1].value)), self._salaries_by_year.items()))

    @property
    def prof_salaries_by_year(self):
        if len(self._prof_salaries_by_year) == 0:
            return {2022: 0}
        return dict(map(lambda item: (item[0], int(item[1].value)), self._prof_salaries_by_year.items()))

    @property
    def prof_counts_by_year(self):
        if len(self._prof_counts_by_year) == 0:
            return {2022: 0}
        return self._prof_counts_by_year

    def _one_percent_filter(self, items_by_cities):
        one_percent = self._vacancies_count / 100
        for city, value in items_by_cities:
            if self._counts_by_cities[city] >= one_percent:
                yield city, value

    @property
    def top_10_salaries_by_cities(self):
        salaries = map(lambda item: (item[0], int(item[1].value)), self._salaries_by_cities.items())
        filtered_items = self._one_percent_filter(salaries)
        return dict(sorted(filtered_items, key=lambda item: -item[1])[:10])

    @property
    def top_10_cities_shares(self):
        result = {}
        for city, count in self._one_percent_filter(self._counts_by_cities.items()):
            result[city] = round(count / self._vacancies_count, 4)
        return dict(sorted(result.items(), key=lambda item: -item[1])[:10])

    def get_top_10_percent_cities_shares(self, digits=None):
        handler = lambda n: 100 * n
        if digits is not None:
            handler = lambda n: round(100 * n, digits)

        return {k: handler(v) for k, v in self.top_10_cities_shares.items()}

    @staticmethod
    def _to_row_dict(title_row, row):
        if len(title_row) != len(row):
            return None
        row_dict = {}
        for i in range(len(row)):
            if row[i] == '':
                return None
            row_dict[title_row[i]] = row[i]
        return row_dict


class ReportColumn:
    def __init__(self, name, values_dict: dict):
        self.name = name
        self._values_dict = values_dict

    @staticmethod
    def with_percents(name, values_dict: dict):
        new_dict = {k: f'{v}%' for k, v in values_dict.items()}
        return ReportColumn(name, new_dict)

    @property
    def keys(self):
        return self._values_dict.keys()

    @property
    def values(self):
        return self._values_dict.values()

    @property
    def pairs(self):
        return self._values_dict.items()


def check_all_equal(iterable, exception):
    g = groupby(iterable)
    all_equal = next(g, True) and not next(g, False)
    if not all_equal:
        raise exception


class ReportSheet:
    def __init__(self, title: str, key_name: str, report_columns: Iterable[ReportColumn], separated=False):
        check_all_equal(map(lambda col: col.keys, report_columns),
                        Exception('Keys of all report_columns must be identical on non-separated sheet'))
        self.title = title
        self.key_name = key_name
        self.report_columns = report_columns
        self.separated = separated

    @property
    def first_column(self):
        return next(iter(self.report_columns))


class ReportTable:
    def __init__(self, bold_titles=True, borders=True):
        self.bold_titles = bold_titles
        self.borders = borders
        self._wb = Workbook()

    def save(self, filename):
        self._wb.save(filename)

    def fill_sheets(self, sheets: Iterable[ReportSheet]):
        first_sheet = True
        for sheet in sheets:
            if first_sheet:
                ws = self._wb.active
                ws.title = sheet.title
                first_sheet = False
            else:
                ws = self._wb.create_sheet(sheet.title)

            if not sheet.separated:
                self._fill_united_sheet(ws, sheet)
            else:
                self._fill_separated_sheet(ws, sheet)
            self._adjust_columns(ws)

    def _fill_united_sheet(self, ws: Worksheet, sheet: ReportSheet):
        self._write_column(ws, title=sheet.key_name, cells=sheet.first_column.keys, column_id=1)
        for column_id, column in enumerate(sheet.report_columns, start=2):
            self._write_column(ws, title=column.name, cells=column.values, column_id=column_id)

    def _fill_separated_sheet(self, ws: Worksheet, sheet: ReportSheet):
        i = 1
        for column in sheet.report_columns:
            self._write_column(ws, title=sheet.key_name, cells=column.keys, column_id=i)
            self._write_column(ws, title=column.name, cells=column.values, column_id=i + 1)
            self._write_empty_column(ws, column_id=i + 2)
            i += 3

    def _format_cell(self, cell):
        bd = Side(style='thin', color="000000")
        border = Border(left=bd, top=bd, right=bd, bottom=bd)
        if self.borders:
            cell.border = border
        return cell

    def _write_column(self, ws: Worksheet, title, cells, column_id):
        title_cell = ws.cell(row=1, column=column_id, value=title)
        self._format_cell(title_cell)
        if self.bold_titles:
            title_cell.font = Font(bold=True)

        for row_id, value in enumerate(cells, start=2):
            cell = ws.cell(row=row_id, column=column_id, value=value)
            self._format_cell(cell)

    @staticmethod
    def _write_empty_column(ws, column_id):
        ws.cell(row=1, column=column_id, value=' ')

    @staticmethod
    def _adjust_columns(ws):
        # Adjusting with max row len + 1 to make numbers visible
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)) + 1))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value


class ReportGraphic:
    def __init__(self):
        plt.rcdefaults()
        self._fig, ((self._ax1, self._ax2), (self._ax3, self._ax4)) = plt.subplots(2, 2, figsize=(8, 8))
        self._fig.tight_layout(pad=5)

    def save_picture(self, filename):
        plt.savefig(filename)

    def get_png_base64_bytes(self):
        io_bytes = io.BytesIO()
        plt.savefig(io_bytes, format='png')
        io_bytes.seek(0)
        return base64.b64encode(io_bytes.read())

    def set_vacancies_statics(self, stats: VacanciesStatistics):
        self._set_year_salaries(stats.salaries_by_year, stats.prof_salaries_by_year)
        self._set_year_counts(stats.counts_by_year, stats.prof_counts_by_year)
        self._set_city_salaries(stats.top_10_salaries_by_cities)
        self._set_city_shares(stats.get_top_10_percent_cities_shares(digits=2))

    def _set_year_salaries(self, general_salaries: dict, prof_salaries: dict):
        labels = list(general_salaries.keys())
        salaries = list(general_salaries.values())
        prof_salaries = list(prof_salaries.values())

        x = np.arange(len(labels))
        width = 0.35  # the width of the bars
        self._ax1.bar(x - width / 2, salaries, width, label='средняя з/п')
        self._ax1.bar(x + width / 2, prof_salaries, width, label='з/п программист')

        self._ax1.set_title('Уровень зарлат по годам')
        self._ax1.set_xticks(x, labels, fontsize=8, rotation=90)
        self._ax1.tick_params(axis='y', labelsize=8)
        self._ax1.legend(fontsize=8)
        self._ax1.yaxis.grid(True)

    def _set_year_counts(self, general_counts: dict, prof_counts: dict):
        labels = list(general_counts.keys())
        counts = list(general_counts.values())
        prof_counts = list(prof_counts.values())

        x = np.arange(len(labels))
        width = 0.35  # the width of the bars
        self._ax2.bar(x - width / 2, counts, width, label='Количество вакансий')
        self._ax2.bar(x + width / 2, prof_counts, width, label='\n'.join(wrap('Количество вакансий программист', 20)))

        self._ax2.set_title('Количество вакансий по годам')
        self._ax2.set_xticks(x, labels, fontsize=8, rotation=90)
        self._ax2.tick_params(axis='y', labelsize=8)
        self._ax2.legend(fontsize=8)
        self._ax2.yaxis.grid(True)

    def _set_city_salaries(self, city_salaries: dict):
        cities = city_salaries.keys()
        salaries = city_salaries.values()

        # Replace spaces and hyphens to \n
        cities = list(map(lambda city: city.replace('-', '-\n').replace(' ', '\n'), cities))
        y_pos = np.arange(len(cities))

        self._ax3.barh(y_pos, salaries, align='center')
        self._ax3.tick_params(axis='y', labelsize=8)
        self._ax3.set_yticks(y_pos, labels=cities, fontsize=6)
        self._ax3.xaxis.grid(True)
        self._ax3.invert_yaxis()  # labels read top-to-bottom
        self._ax3.set_title('Уровень зарплат по городам')

    def _set_city_shares(self, city_shares: dict):
        raw_labels = list(city_shares.keys())
        raw_sizes = list(city_shares.values())

        labels = ['Другие'] + raw_labels
        sizes = [100 - sum(raw_sizes)] + raw_sizes

        self._ax4.set_title('Доля вакансий по городам')
        self._ax4.pie(sizes, labels=labels,
                      shadow=False, startangle=0, textprops={'fontsize': 6})
        self._ax4.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.


class Report:
    def __init__(self, stats: VacanciesStatistics):
        self.stats = stats

    def print(self):
        print('Динамика уровня зарплат по годам:', self.stats.salaries_by_year)
        print('Динамика количества вакансий по годам:', self.stats.counts_by_year)
        print('Динамика уровня зарплат по годам для выбранной профессии:', self.stats.prof_salaries_by_year)
        print('Динамика количества вакансий по годам для выбранной профессии:', self.stats.prof_counts_by_year)
        print('Уровень зарплат по городам (в порядке убывания):', self.stats.top_10_salaries_by_cities)
        print('Доля вакансий по городам (в порядке убывания):', self.stats.top_10_cities_shares)

    def generate_excel(self, filename, bold_titles=True, borders=True):
        year_columns = []
        year_columns += [ReportColumn('Средняя зарпалата', self.stats.salaries_by_year)]
        year_columns += [ReportColumn(f'Средняя зарпалата - {prof_name}', self.stats.prof_salaries_by_year)]
        year_columns += [ReportColumn(f'Количество вакансий', self.stats.counts_by_year)]
        year_columns += [ReportColumn(f'Количество вакансий - {prof_name}', self.stats.prof_counts_by_year)]
        year_sheet = ReportSheet('Статистика по годам', 'Год', year_columns)
        city_columns = []
        city_columns += [ReportColumn('Уровень зарплат', self.stats.top_10_salaries_by_cities)]
        percent_shares = self.stats.get_top_10_percent_cities_shares(digits=2)
        city_columns += [ReportColumn.with_percents('Доля вакансий', percent_shares)]
        city_sheet = ReportSheet('Статистика по городам', 'Город', city_columns, separated=True)

        table = ReportTable(bold_titles, borders)
        table.fill_sheets([year_sheet, city_sheet])
        table.save(filename)

    def generate_image(self, filename):
        graphic = ReportGraphic()
        graphic.set_vacancies_statics(self.stats)
        graphic.save_picture(filename)

    def _get_base64_png(self):
        graphic = ReportGraphic()
        graphic.set_vacancies_statics(self.stats)
        return graphic.get_png_base64_bytes().decode("utf-8")

    def generate_pdf(self, filename):
        def get_year_items():
            years = self.stats.salaries_by_year.keys()
            salaries = self.stats.salaries_by_year.values()
            prof_salaries = self.stats.prof_salaries_by_year.values()
            counts = self.stats.counts_by_year.values()
            prof_counts = self.stats.prof_counts_by_year.values()
            for item in zip(years, salaries, prof_salaries, counts, prof_counts):
                yield {
                    'year': item[0],
                    'salary': item[1],
                    'prof_salary': item[2],
                    'counts': item[3],
                    'prof_counts': item[4],
                }

        def get_city_salaries():
            city_salaries = self.stats.top_10_salaries_by_cities
            for city, salary in city_salaries.items():
                yield {
                    'city': city,
                    'salary': salary
                }

        def get_city_shares():
            city_shares = self.stats.get_top_10_percent_cities_shares(digits=2)
            for city, share_val in city_shares.items():
                yield {
                    'city': city,
                    'share': f'{share_val}%'
                }

        env = Environment(loader=FileSystemLoader('..'))
        template = env.get_template('report_template.html')
        html = template.render({
            'prof_name': 'Программист',
            'graph_bytes': self._get_base64_png()
        },
            year_items=get_year_items(),
            city_salaries=get_city_salaries(),
            city_shares=get_city_shares(),
        )
        config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)
        pdfkit.from_string(html, filename, configuration=config)


def execute_program(file_name, prof_name):
    stats = VacanciesStatistics(file_name, prof_name)
    report = Report(stats)
    print('Формирование таблицы...')
    report.generate_excel('report.csv')
    print('Файл с таблицей готов!')
    print('Формирование файла c графиками...')
    report.generate_image('graph.png')
    print('Изображение графика готово!')
    print('Формирование PDF...')
    report.generate_pdf('report.pdf')
    print('PDF готов!')


if __name__ == "__main__":
    file_name = input('Введите название файла: ')
    prof_name = input('Введите название профессии: ')
    execute_program(file_name, prof_name)
