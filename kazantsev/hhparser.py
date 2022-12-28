import base64
import csv
import io
import itertools
import os
import re
from itertools import groupby
from textwrap import wrap
from typing import Iterable

import matplotlib
import numpy as np
import pdfkit
from jinja2 import Environment, FileSystemLoader
from matplotlib import pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Side, Border
from openpyxl.worksheet.worksheet import Worksheet
from prettytable import PrettyTable, ALL

from kazantsev.local_path import get_local_path
from kazantsev.models import Salary, Vacancy
from kazantsev.vacancies_staticstics import VacanciesStatistics, SingleProcessVacanciesStatistics

matplotlib.use('TkAgg')

WKHTMLTOPDF_PATH = r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe'
HTML_TEMPLATE_PATH = r'report_template.html'

exp_naming = {
    "noExperience": "Нет опыта",
    "between1And3": "От 1 года до 3 лет",
    "between3And6": "От 3 до 6 лет",
    "moreThan6": "Более 6 лет"
}

cur_naming = {
    "AZN": "Манаты",
    "BYR": "Белорусские рубли",
    "EUR": "Евро",
    "GEL": "Грузинский лари",
    "KGS": "Киргизский сом",
    "KZT": "Тенге",
    "RUR": "Рубли",
    "UAH": "Гривны",
    "USD": "Доллары",
    "UZS": "Узбекский сум"
}

gross_naming = {
    False: 'С вычетом налогов',
    True: 'Без вычета налогов'
}

bool_naming = {
    True: 'Да',
    False: 'Нет'
}

reversed_bool_naming = {
    'Да': True,
    'Нет': False
}


def format_sum(sum_value):
    """
    Преобразовываем указанное число в строчку, где у числа тысячи разделены пробелами
    :param sum_value: Сумма
    :type sum_value: float or int or str
    :return: строчку с данным числом, у которого тысячи разделены пробелом
    :rtype: str

    >>> format_sum(50000)
    '50 000'
    >>> format_sum(77888.77)
    '77 888'
    >>> format_sum('123456789')
    '123 456 789'
    """
    return format(int(float(sum_value)), ',').replace(',', ' ')


class EmptyFileException(Exception):
    """
    Исключение, выбрасываемое в случае пустого обрабатываемого файла
    """
    pass


NO_TAG_PATTERN = re.compile(r"<.*?>")
SINGLE_SPACES_PATTERN = re.compile(r"\s+")


class DataSet:
    """
    Класс, который обрабатывает файл с вакансиями и дает возможность получить их в виде объектов

    Attributes:
        file_name (str): Путь к файлу в локальной системе
        vacancy_objects (List[Vacancy]): Список вакансий
    """

    def __init__(self, file_name):
        """
        Инициализирует экземляр класса DataSet
        :param file_name: Путь к файлу в локальной системе
        :type file_name: str
        :raises EmptyFileException: Если обрабатываемый файл пуст
        """
        self.file_name = file_name

        # Read vacancies from csv file
        list_naming, rows = self._read_csv(file_name)
        if list_naming is None:
            raise EmptyFileException()
        row_dicts = self._get_row_dicts(rows, list_naming)
        self.vacancy_objects = self._get_vacancy_objects(row_dicts)

    @staticmethod
    def _transform_value(value: str):
        """
        Преобразует сырое значение из таблицы в список или строку (если нет \n в значении) с очищенными тегами и
        пробелами
        :param value: Сырое значение
        :type value: str
        :return: Очищеное значение в виде str или list, есть обработано несколько значение, разделенных \n
        :rtype: str or List[str]
        """
        result = []
        for sub_value in value.split('\n'):
            no_tag_value = NO_TAG_PATTERN.sub('', sub_value)
            clean_value = SINGLE_SPACES_PATTERN.sub(' ', no_tag_value.strip())
            result.append(clean_value)
        if len(result) == 1:
            return result[0]
        else:
            return result

    # Returns header fields and list of other csv fields:
    @staticmethod
    def _read_csv(file_name):
        """
        Считывает заголовочную строку и остальные из csv файла
        :param file_name:
        :type file_name: str
        :return: Кортеж из списка заголовков и списка строк
        :rtype: Tuple[List[str], List[List[str]]]
        """
        absolute = get_local_path(file_name).absolute()
        with open(absolute, 'r', encoding='utf_8_sig') as file:
            reader = csv.reader(file)
            rows = [row for row in reader]
            if len(rows) == 0 or all(len(row) == 0 or (len(row) == 1 and row[0].strip() == '') for row in rows):
                return None, None
            list_naming = rows.pop(0)
            return list_naming, rows

    @staticmethod
    def _get_row_dicts(rows, list_naming):
        """
        Генерирует из обработанных строк и заголовочной строки словари
        :param rows: Обработанные строки
        :type rows: List[List[str]]
        :param list_naming: Заголовочная строка
        :type list_naming: List[str]
        :return: Итератор словарей
        :rtype: Iterator[Dict[str, str]]
        """
        for row_values in rows:
            row_dict = {}
            if len(row_values) != len(list_naming) or '' in row_values:
                continue
            for i in range(len(row_values)):
                row_dict[list_naming[i]] = DataSet._transform_value(row_values[i])
            yield row_dict

    @staticmethod
    def _get_vacancy_objects(row_dicts):
        """
        Преобразует словари в объекты вакансий
        :param row_dicts: Словари
        :return: Список объектов вакансий
        :rtype: List[Vacancy]
        """
        objects = []
        for row_dict in row_dicts:
            vacancy = Vacancy.parse_from_dict(row_dict)
            objects.append(vacancy)
        return objects


class FilterException(Exception):
    """
    Исключение, связанное с фильтровкой вакансий
    """
    pass


class SortException(Exception):
    """
    Исключение, связанное с сортировкой вакансий
    """
    pass


class NumberStringException(Exception):
    """
    Исключение, свзанное со строкой ограничения по номерам и количеству
    """
    pass


class InputConect:
    """
    Класс, который принимает данные о вакансий, параметры и позволяет выводить их в виде таблицы в строке

    Attributes:
        filter_key (str): Ключ для фильтрации
        filter_value (str): Значение для фильтрации
        sorter (Callable[[Vacancy], any]): Делегат, выдающий ключ для сортировки
        reverse_sort (bool): Перевернута ли сортировка
        start_number (int or None): Номер вакансии, с которой будет начинаться диапазон (отсчет вакансий с 1)
        end_number (int or None): Номер вакансии, которой будет окончен диапазон (она не будет включена в него,
        отсчет вакансий с 1)
        columns_string (str): Строка с названиями столбцов через запятую и пробел
        columns_naming (Dict[str, str]): Словарь кодовое название => русское название
    """

    def __init__(self):
        """
        Инициализирует объект InputConect
        См. методы для установки параметров для вывода таблицы
        """
        self._filters = {
            'Название': lambda vac, exp_val: vac.name == exp_val,
            'Описание': lambda vac, exp_val: vac.description == exp_val,
            'Компания': lambda vac, exp_val: vac.employer_name == exp_val,
            'Название региона': lambda vac, exp_val: vac.area_name == exp_val,
            'Оклад': lambda vac, exp_val: vac.salary.from_amount <= int(exp_val) <= vac.salary.to_amount,
            'Опыт работы': lambda vac, exp_val: exp_naming[vac.experience_id] == exp_val,
            'Премиум-вакансия': lambda vac, exp_val: bool_naming[vac.premium] == exp_val,
            'Идентификатор валюты оклада': lambda vac, exp_val: cur_naming[vac.salary.currency] == exp_val,
            'Дата публикации вакансии': lambda vac, exp_val: vac.str_publish_date == exp_val,
            'Навыки': lambda vac, exp_val: all([s in vac.key_skills for s in exp_val.split(', ')]),
        }
        self._sorters = {
            'Название': lambda vac: vac.name,
            'Описание': lambda vac: vac.description,
            'Компания': lambda vac: vac.employer_name,
            'Название региона': lambda vac: vac.area_name,
            'Навыки': lambda vac: len(vac.key_skills),
            'Оклад': lambda vac: vac.salary.avg_ruble_amount,
            'Дата публикации вакансии': lambda vac: vac.published_at,
            'Опыт работы': lambda vac: list(exp_naming.keys()).index(vac.experience_id),
            'Премиум-вакансия': lambda vac: int(vac.premium),
            'Идентификатор валюты оклада': lambda vac: vac.salary.currency,
        }

        self.filter_key = None
        self.filter_value = None
        self.sorter = None
        self.reverse_sort = False
        self.start_number = None
        self.end_number = None
        self.columns_string = None
        self.columns_naming = {
            'name': 'Название',
            'description': 'Описание',
            'key_skills': 'Навыки',
            'experience_id': 'Опыт работы',
            'premium': 'Премиум-вакансия',
            'employer_name': 'Компания',
            'salary': 'Оклад',
            'area_name': 'Название региона',
            'published_at': 'Дата публикации вакансии'
        }

    def set_filter_string(self, filter_string):
        """
        Установка строки для фильтрации
        :param filter_string: Строка формата "Ключ: НужноеЗначение"
        :return: None
        """
        if filter_string == '':
            return

        args = filter_string.split(': ')
        if len(args) != 2:
            raise FilterException('Формат ввода некорректен')
        if args[0] not in self._filters:
            raise FilterException('Параметр поиска некорректен')

        self.filter_key = args[0]
        self.filter_value = args[1]

    def set_sort_params(self, sort_string, reverse_str: str):
        """
        Устновка параметров для фильтрации
        :param sort_string: Название столбца на русском языке с большой буквы
        :type sort_string: str
        :param reverse_str: Перевернут ли порядок сортировка
        :type reverse_str: str
        :return: None
        """
        if sort_string == '':
            pass
        elif sort_string not in self._sorters:
            raise SortException('Параметр сортировки некорректен')
        else:
            self.sorter = self._sorters[sort_string]

        if reverse_str == '':
            pass
        elif reverse_str not in bool_naming.values():
            raise SortException('Порядок сортировки задан некорректно')
        else:
            self.reverse_sort = reversed_bool_naming[reverse_str]

    def set_number_string(self, number_string):
        """
        Устновка параметров для установки диапазона
        :param number_string: Строка формата 'начало' или 'начало конец',
        :type number_string: str
        :return: None
        :raises NumberStringException: Если строка содержит не 1-2 числа
        :raises ValueError: Если не удалось спарсить числа
        """
        if number_string == "":
            return

        values = number_string.strip().split(' ')
        if len(values) == 1:
            self.start_number = int(values[0])
        elif len(values) == 2:
            self.start_number = int(values[0])
            self.end_number = int(values[1])
        else:
            raise NumberStringException('Диапазон вакансий должен содержать 1 или 2 числа.')

    def set_columns_string(self, columns_string):
        """
        Устанавливает строку с названиями столбцов
        :param columns_string: Строка с отображаемыми названиями столбцов через запятую и пробел
        :type columns_string: str
        :return: None
        """
        if columns_string == "":
            return
        self.columns_string = columns_string

    def print_as_table(self, vacancies):
        """
        Печатает вакансии в виде таблицы в консоли
        :param vacancies: Список вакансий
        :type vacancies: List[Vacancy]
        :return: None
        :raises FilterException: Если не найдено нужных по фильтрам строк.
        """
        if self.filter_key is not None:
            vacancies = self._filter_vacancies_by_string(vacancies)

        if self.sorter is not None:
            vacancies = sorted(vacancies, key=self.sorter, reverse=self.reverse_sort)

        table = self._create_table(vacancies)
        if table is None:
            print('Нет данных')
            return

        table_formatter = self._get_table_formatter()
        print(table_formatter(table))

    def _create_table(self, vacancies_objects):
        """
        Создает объект таблицы или выдает None если вакансий ноль
        :param vacancies_objects: Список вакансий
        :type vacancies_objects: List[Vacancy]
        :return: Объект таблицы
        :rtype: PrettyTable
        """
        table = PrettyTable()
        table.hrules = ALL
        table.field_names = ['№'] + list(self.columns_naming.values())
        table.align = 'l'
        table.max_width = 20

        n = 1
        for vacancy in vacancies_objects:
            formatted_vacancy = self._format_vacancy(vacancy)
            table.add_row([str(n)] + list(map(self._cut_string, formatted_vacancy.values())))
            n += 1

        if n <= 1:
            return None
        else:
            return table

    @staticmethod
    def _cut_string(value):
        """
        Режет строку при 100 и более символов в ней
        :param value: Обрабатываемая строка
        :type value: str
        :return: Готовая строка
        :rtype: str
        """
        if len(value) > 100:
            return value[:100] + '...'
        else:
            return value

    @staticmethod
    def _format_vacancy(vacancy: Vacancy):
        """
        Преобразует вакансию в словарь для таблицы
        :param vacancy: Вакансия
        :type: Vacancy
        :return: Словарь
        :rtype: Dict[str, any]
        """
        new_row = dict()
        # Форматируем название и описание
        new_row['name'] = vacancy.name
        new_row['description'] = vacancy.description
        new_row['key_skills'] = '\n'.join(vacancy.key_skills)
        # Заменяем опыт работы
        new_row['experience_id'] = exp_naming[vacancy.experience_id]
        # Руссифицируем bool
        new_row['premium'] = bool_naming[vacancy.premium]
        # Получаем имя работодателя
        new_row['employer_name'] = vacancy.employer_name
        # Меняем оклад
        new_row['salary'] = InputConect._format_salary(vacancy.salary)
        # Получаем регион
        new_row['area_name'] = vacancy.area_name
        # Заменяем дату публикации
        new_row['published_at'] = vacancy.str_publish_date
        # Возвращаем словарь с форматированными значениями
        return new_row

    @staticmethod
    def _format_salary(salary: Salary):
        """
        Преобразует зарплату в строку
        :param salary: Объект зарплаты
        :type salary: Salary
        :return: Строка формата 'окладОт - окладДо (валюта) (сВычетомНалоговИлиБез)'
        :rtype: str
        """
        sal_from = format_sum(salary.from_amount)
        sal_to = format_sum(salary.to_amount)
        sal_gross = gross_naming[salary.gross]
        sal_currency_id = salary.currency
        sal_currency = cur_naming[sal_currency_id]
        return f'{sal_from} - {sal_to} ({sal_currency}) ({sal_gross})'

    def _get_table_formatter(self):
        """
        Служебный метод для таблицы
        :return: Лямбда для обработки значений
        :rtype: Callable
        """
        formatter_args = dict()
        if self.start_number is not None:
            formatter_args['start'] = self.start_number - 1
        if self.end_number is not None:
            formatter_args['end'] = self.end_number - 1
        if self.columns_string is not None:
            formatter_args['fields'] = ['№'] + self.columns_string.split(', ')
        return lambda t: t.get_string(**formatter_args)

    def _filter_vacancies_by_string(self, vacancies):
        """
        Служебный метод для фильтрации вакансий по self.filter_key и self.filter_value
        :param vacancies: Вакансии
        :type vacancies: List[Vacancy]
        :return: Отфильтрованыые вакансии
        :rtype: Iterable[Vacancy]
        :raises FilterException: Если не найдено нужных по фильтрам строк
        """

        def peek(iterable):
            try:
                first = next(iterable)
            except StopIteration:
                return None
            return first, itertools.chain([first], iterable)

        filtered = filter(lambda vac: self._filters[self.filter_key](vac, self.filter_value), vacancies)
        filtered_peek = peek(filtered)
        if filtered_peek is None:
            raise FilterException('Ничего не найдено')
        return filtered_peek[1]


class ReportColumn:
    """
    Класс колонки данных в репорте

    Attributes:
        name (str): Название колонки
    """

    def __init__(self, name, values_dict: dict):
        """
        Инициализация
        :param name: Назавние колонки
        :type name: str
        :param values_dict: Словарь ключей и значений для колонки
        :type values_dict: Dict[str, str]
        """
        self.name = name
        self._values_dict = values_dict

    @staticmethod
    def with_percents(name, values_dict: dict):
        """
        Инициализация значений, к которым приписывается процент
        :param name: Назавние колонки
        :type name: str
        :param values_dict: Словарь ключей и значений для колонки
        :type values_dict: Dict[str, float]
        """
        new_dict = {k: f'{v}%' for k, v in values_dict.items()}
        return ReportColumn(name, new_dict)

    @property
    def keys(self):
        """
        Возвращает ключи
        :return: Ключи
        :rtype: Iterable[any]
        """
        return self._values_dict.keys()

    @property
    def values(self):
        """
        Возвращает значения
        :return: Значения
        :rtype: Iterable[any]
        """
        return self._values_dict.values()


def check_all_equal(iterable, exception):
    """
    Проверяет, совпадают ли все значения в итераторе, иначе выдают исключение
    :param iterable: Итератор
    :type iterable: Iterable
    :param exception: Исключение
    :type exception: Exception
    :return: None
    """
    g = groupby(iterable)
    all_equal = next(g, True) and not next(g, False)
    if not all_equal:
        raise exception


class ReportSheet:
    """
    Класс листа таблицы

    Attributes:
        title (str): Название листа
        key_name (str): Название ключевых значений
        report_columns (Iterable[ReportColumn]): Колонки для данного листа
        separated (bool): Разделены ли колонки на листе
    """

    def __init__(self, title: str, key_name: str, report_columns: Iterable[ReportColumn], separated=False):
        """
        Инициализацирует лист таблицы
        :param title: Название листа
        :type title: str
        :param key_name: Название ключевых значений
        :type key_name: str
        :param report_columns: Колонки для данного листа
        :type report_columns: Iterable[ReportColumn]
        :param separated: Разделена ли таблица
        :type separated: bool
        """
        if not separated:
            check_all_equal(map(lambda col: col.keys, report_columns),
                            Exception('Keys of all report_columns must be identical on non-separated sheet'))
        self.title = title
        self.key_name = key_name
        self.report_columns = report_columns
        self.separated = separated

    @property
    def first_column(self):
        """
        Возвращает первую колонку в листе
        :return: Первая колонка
        :rtype: ReportColumn
        """
        return next(iter(self.report_columns))


class ReportTable:
    """
    Класс таблицы репорта

    Attributes:
        bold_titles (bool): Жирные ли заголовки
        borders (bool): Есть ли рамки для ячеек
    """

    def __init__(self, bold_titles=True, borders=True):
        """
        Инициализирует объект таблицы репорта
        :param bold_titles: Жирные ли заголовки
        :param borders: Есть ли рамки для ячеек
        :type bold_titles: bool
        :type borders: bool
        """
        self.bold_titles = bold_titles
        self.borders = borders
        self._wb = Workbook()

    def save(self, filename):
        """
        Сохраняет таблицу в файл по указанному пути
        :param filename: Путь к таблице
        :type filename: str
        :return: None
        """
        self._wb.save(filename)

    def fill_sheets(self, sheets: Iterable[ReportSheet]):
        """
        Заполняет листы в таблице
        :param sheets: Объекты листов
        :type sheets: Iterable[ReportSheets]
        :return: None
        """
        first_sheet = True
        for sheet in sheets:
            if first_sheet:
                ws = self._wb.active
                ws.title = sheet.title
                first_sheet = False
            else:
                ws = self._wb.create_sheet(sheet.title)

            if not sheet.separated:
                self._fill_united_sheet(ws, sheet)
            else:
                self._fill_separated_sheet(ws, sheet)
            self._adjust_columns(ws)

    def _fill_united_sheet(self, ws: Worksheet, sheet: ReportSheet):
        """
        Заполняет лист с объеденнеными колонками
        :param ws: Рабочая таблица
        :param sheet: Лист таблицы
        :type ws: Worksheet
        :type sheet: ReportSheet
        :return: None
        """
        self._write_column(ws, title=sheet.key_name, cells=sheet.first_column.keys, column_id=1)
        for column_id, column in enumerate(sheet.report_columns, start=2):
            self._write_column(ws, title=column.name, cells=column.values, column_id=column_id)

    def _fill_separated_sheet(self, ws: Worksheet, sheet: ReportSheet):
        """
        Заполняет лист с разделенными колонками
        :param ws: Рабочая таблица
        :param sheet: Лист таблицы
        :type ws: Worksheet
        :type sheet: ReportSheet
        :return: None
        """
        i = 1
        for column in sheet.report_columns:
            self._write_column(ws, title=sheet.key_name, cells=column.keys, column_id=i)
            self._write_column(ws, title=column.name, cells=column.values, column_id=i + 1)
            self._write_empty_column(ws, column_id=i + 2)
            i += 3

    def _format_cell(self, cell):
        """
        Форматирует объект ячейки
        :param cell: Объект ячейки
        :return: Форматированный объект ячейки
        :type cell: Cell
        :rtype: Cell
        """
        bd = Side(style='thin', color="000000")
        border = Border(left=bd, top=bd, right=bd, bottom=bd)
        if self.borders:
            cell.border = border
        return cell

    def _write_column(self, ws: Worksheet, title, cells, column_id):
        """
        Записывает колонку в объект рабочей таблицы
        :param ws: Рабочая таблица
        :param title: Заголовок колонки
        :param cells: Ячейки
        :param column_id: id колонки (начиная с 1)
        :return: None
        :type ws: Worksheet
        :type title: str
        :type cells: Iterable[Cell]
        :type column_id: int
        """
        title_cell = ws.cell(row=1, column=column_id, value=title)
        self._format_cell(title_cell)
        if self.bold_titles:
            title_cell.font = Font(bold=True)

        for row_id, value in enumerate(cells, start=2):
            cell = ws.cell(row=row_id, column=column_id, value=value)
            self._format_cell(cell)

    @staticmethod
    def _write_empty_column(ws, column_id):
        """
        Записывает пустую колонку в рабочую таблицу
        :param ws: Рабочая таблица
        :param column_id: id колонки (начиная с 1)
        :type ws: Worksheet
        :type column_id: int
        :return: None
        """
        ws.cell(row=1, column=column_id, value=' ')

    @staticmethod
    def _adjust_columns(ws):
        """
        Адаптирует все колонки в рабочей таблицы по размеру, с которым они будут видимыми
        :param ws: Рабочая таблица
        :return: None
        :type ws: Worksheet
        """
        # Adjusting with max row len + 1 to make numbers visible
        dims = {}
        for row in ws.rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)) + 1))
        for col, value in dims.items():
            ws.column_dimensions[col].width = value


class ReportGraphic:
    """
    Класс для работы с графикой
    """

    def __init__(self, prof_name):
        """
        Инициализация объекта для работы с графикой
        """
        self.prof_name = prof_name
        plt.rcdefaults()
        self._fig, ((self._ax1, self._ax2), (self._ax3, self._ax4)) = plt.subplots(2, 2, figsize=(8, 8))
        self._fig.tight_layout(pad=5)

    def save_picture(self, filename):
        """
        Сохраняет график в виде изображения
        :param filename: Путь к файлу
        :return: None
        :type filename: str
        """
        plt.savefig(filename)

    def get_png_base64_bytes(self):
        """
        Возвращает изображения график в виде байтов base64
        :return: Байты base64
        :rtype: bytes
        """
        io_bytes = io.BytesIO()
        plt.savefig(io_bytes, format='png')
        io_bytes.seek(0)
        return base64.b64encode(io_bytes.read())

    def set_vacancies_statics(self, stats: VacanciesStatistics):
        """
        Устанавливает объект статистики по вакансиям для построения графиков
        :param stats: Статистика
        :return: None
        :type stats: VacanciesStatistics
        """
        self._set_year_salaries(stats.salaries_by_year, stats.prof_salaries_by_year)
        self._set_year_counts(stats.counts_by_year, stats.prof_counts_by_year)
        self._set_city_salaries(stats.top_10_salaries_by_cities)
        self._set_city_shares(stats.get_top_10_percent_cities_shares(digits=2))

    def _set_year_salaries(self, general_salaries: dict, prof_salaries: dict):
        """
        Рисует часть графика со статистикой зарплат по годам
        :param general_salaries: Все зарплаты по годам
        :param prof_salaries: Зарплаты для данной профессии по годам
        :return: None
        :type general_salaries: dict
        :type prof_salaries: dict
        """
        labels = list(general_salaries.keys())
        salaries = list(general_salaries.values())
        prof_salaries = list(prof_salaries.values())

        x = np.arange(len(labels))
        width = 0.35  # the width of the bars
        self._ax1.bar(x - width / 2, salaries, width, label='средняя з/п')
        self._ax1.bar(x + width / 2, prof_salaries, width, label=f'з/п {self.prof_name}')

        self._ax1.set_title('Уровень зарлат по годам')
        self._ax1.set_xticks(x, labels, fontsize=8, rotation=90)
        self._ax1.tick_params(axis='y', labelsize=8)
        self._ax1.legend(fontsize=8)
        self._ax1.yaxis.grid(True)

    def _set_year_counts(self, general_counts: dict, prof_counts: dict):
        """
        Рисует часть графика со статистикой количества вакансий по годам
        :param general_counts: Все количества по годам
        :param prof_counts: Количества вакансий для данной профессии по годам
        :return: None
        :type general_counts: dict
        :type prof_counts: dict
        """
        labels = list(general_counts.keys())
        counts = list(general_counts.values())
        prof_counts = list(prof_counts.values())

        x = np.arange(len(labels))
        width = 0.35  # the width of the bars
        self._ax2.bar(x - width / 2, counts, width, label='Количество вакансий')
        self._ax2.bar(x + width / 2, prof_counts, width, label='\n'.join(wrap(f'Количество вакансий {self.prof_name}', 20)))

        self._ax2.set_title('Количество вакансий по годам')
        self._ax2.set_xticks(x, labels, fontsize=8, rotation=90)
        self._ax2.tick_params(axis='y', labelsize=8)
        self._ax2.legend(fontsize=8)
        self._ax2.yaxis.grid(True)

    def _set_city_salaries(self, city_salaries: dict):
        """
        Рисует часть графика с заралтами по городам
        :param city_salaries: Зарплаты по городам
        :return: None
        :type city_salaries: dict
        """
        cities = city_salaries.keys()
        salaries = city_salaries.values()

        # Replace spaces and hyphens to \n
        cities = list(map(lambda city: city.replace('-', '-\n').replace(' ', '\n'), cities))
        y_pos = np.arange(len(cities))

        self._ax3.barh(y_pos, salaries, align='center')
        self._ax3.tick_params(axis='y', labelsize=8)
        self._ax3.set_yticks(y_pos, labels=cities, fontsize=6)
        self._ax3.xaxis.grid(True)
        self._ax3.invert_yaxis()  # labels read top-to-bottom
        self._ax3.set_title('Уровень зарплат по городам')

    def _set_city_shares(self, city_shares: dict):
        """
        Рисует часть графика с долям по городам
        :param city_shares: Доли по городам
        :return: None
        :type city_shares: dict
        """
        raw_labels = list(city_shares.keys())
        raw_sizes = list(city_shares.values())

        labels = ['Другие'] + raw_labels
        sizes = [100 - sum(raw_sizes)] + raw_sizes

        self._ax4.set_title('Доля вакансий по городам')
        self._ax4.pie(sizes, labels=labels,
                      shadow=False, startangle=0, textprops={'fontsize': 6})
        self._ax4.axis('equal')  # Equal aspect ratio ensures that pie is drawn as a circle.


class Report:
    """
    Класс репорта

    Attributes:
        stats (VacancyStatistics): Статистика вакансий
    """

    def __init__(self, stats: VacanciesStatistics):
        """
        Инициализация
        :param stats: Статистика вакансий
        :type stats: VacanciesStatistics
        """
        self.stats = stats

    def print(self):
        """
        Печатает репорт в консоль
        :return: None
        """
        print('Динамика уровня зарплат по годам:', dict(sorted(self.stats.salaries_by_year.items())))
        print('Динамика количества вакансий по годам:', dict(sorted(self.stats.counts_by_year.items())))
        print('Динамика уровня зарплат по годам для выбранной профессии:',
              dict(sorted(self.stats.prof_salaries_by_year.items())))
        print('Динамика количества вакансий по годам для выбранной профессии:',
              dict(sorted(self.stats.prof_counts_by_year.items())))
        print('Уровень зарплат по городам (в порядке убывания):', self.stats.top_10_salaries_by_cities)
        print('Доля вакансий по городам (в порядке убывания):', self.stats.top_10_cities_shares)

    def generate_excel(self, filename, bold_titles=True, borders=True):
        """
        Генерирует excel файл
        :param filename: Путь к файлу с таблицей
        :param bold_titles: Жирные заголовки
        :param borders: Есть ли рамки у ячеек
        :return: None
        :type filename: str
        :type bold_titles: bool
        :type borders: bool
        """
        year_columns = []
        year_columns += [ReportColumn('Средняя зарпалата', self.stats.salaries_by_year)]
        year_columns += [ReportColumn(f'Средняя зарпалата - {self.stats.prof_name}', self.stats.prof_salaries_by_year)]
        year_columns += [ReportColumn(f'Количество вакансий', self.stats.counts_by_year)]
        year_columns += [ReportColumn(f'Количество вакансий - {self.stats.prof_name}', self.stats.prof_counts_by_year)]
        year_sheet = ReportSheet('Статистика по годам', 'Год', year_columns)
        city_columns = []
        city_columns += [ReportColumn('Уровень зарплат', self.stats.top_10_salaries_by_cities)]
        percent_shares = self.stats.get_top_10_percent_cities_shares(digits=2)
        city_columns += [ReportColumn.with_percents('Доля вакансий', percent_shares)]
        city_sheet = ReportSheet('Статистика по городам', 'Город', city_columns, separated=True)

        table = ReportTable(bold_titles, borders)
        table.fill_sheets([year_sheet, city_sheet])
        table.save(filename)

    def generate_image(self, filename):
        """
        Генерирует изображения графика и сохраняет его
        :param filename: Путь к файлу с изображением
        :return: None
        :type filename: str
        """
        graphic = ReportGraphic(self.stats.prof_name)
        graphic.set_vacancies_statics(self.stats)
        graphic.save_picture(filename)

    def _get_base64_png(self):
        """
        Получает base64 строку графика
        :return: base64 изображение в виде utf-8 строки
        :rtype: str
        """
        graphic = ReportGraphic(self.stats.prof_name)
        graphic.set_vacancies_statics(self.stats)
        return graphic.get_png_base64_bytes().decode("utf-8")

    def generate_pdf(self, filename):
        """
        Генерирует PDF файл со статистикой
        :param filename: Путь к PDF файлу
        :return: None
        :type filename: str
        """

        def get_year_items():
            years = self.stats.salaries_by_year.keys()
            salaries = self.stats.salaries_by_year.values()
            prof_salaries = self.stats.prof_salaries_by_year.values()
            counts = self.stats.counts_by_year.values()
            prof_counts = self.stats.prof_counts_by_year.values()
            for item in zip(years, salaries, prof_salaries, counts, prof_counts):
                yield {
                    'year': item[0],
                    'salary': item[1],
                    'prof_salary': item[2],
                    'counts': item[3],
                    'prof_counts': item[4],
                }

        def get_city_salaries():
            city_salaries = self.stats.top_10_salaries_by_cities
            for city, salary in city_salaries.items():
                yield {
                    'city': city,
                    'salary': salary
                }

        def get_city_shares():
            city_shares = self.stats.get_top_10_percent_cities_shares(digits=2)
            for city, share_val in city_shares.items():
                yield {
                    'city': city,
                    'share': f'{share_val}%'
                }

        env = Environment(loader=FileSystemLoader(os.path.join(os.path.dirname(__file__))))
        template = env.get_template(HTML_TEMPLATE_PATH)

        html = template.render({
            'prof_name': self.stats.prof_name,
            'area_name': self.stats.area_name,
            'graph_bytes': self._get_base64_png()
        },
            year_items=get_year_items(),
            city_salaries=get_city_salaries(),
            city_shares=get_city_shares(),
        )
        config = pdfkit.configuration(wkhtmltopdf=WKHTMLTOPDF_PATH)
        pdfkit.from_string(html, filename, configuration=config)


def execute_vacancies(file_name, filter_string, sort_column, sort_reversed_string, number_string, columns_string):
    """
    Выводит таблицу вакансий в консоль по пользовательским данным
    :param file_name: Путь к файлу с таблицей вакансий
    :param filter_string: Строка для фильтрации формата "Ключ: НужноеЗначение"
    :param sort_column: Название столбца, по которому сортируются вакансии, на русском языке с большой буквы
    :param sort_reversed_string: Перевернут ли порядок сортировки (Да/Нет)
    :param number_string: Строка из чисел формата 'начало' или 'начало конец' или 'начало конец количество'
    :param columns_string: Строка с отображаемыми названиями столбцов через запятую и пробел
    :return: None
    :type file_name: str
    :type filter_string: str
    :type sort_column: str
    :type sort_reversed_string: str
    :type number_string: str
    :type columns_string: str
    """
    conect = InputConect()
    try:
        conect.set_filter_string(filter_string)
        conect.set_sort_params(sort_column, sort_reversed_string)
        conect.set_number_string(number_string)
    except (FilterException, SortException, NumberStringException) as ex:
        print(ex)
        return
    conect.set_columns_string(columns_string)

    try:
        data = DataSet(file_name)
    except EmptyFileException:
        print('Пустой файл')
        return

    try:
        conect.print_as_table(data.vacancy_objects)
    except FilterException as ex:
        print(ex)


def execute_reports(file_name, prof_name):
    """
    Формирует репорты со стастикой по вакансиям в виде таблицы, изображения и PDF-файла
    :param file_name: Путь к файлу с таблицей с вакансиями
    :param prof_name: Название профессий
    :type file_name: str
    :type prof_name: str
    :return: None
    """
    # TODO: Change class
    stats = SingleProcessVacanciesStatistics(file_name, prof_name)
    report = Report(stats)
    print('Формирование таблицы...')
    report.generate_excel('report.csv')
    print('Файл с таблицей готов!')
    print('Формирование файла c графиками...')
    report.generate_image('graph.png')
    print('Изображение графика готово!')
    print('Формирование PDF...')
    report.generate_pdf('report.pdf')
    print('PDF готов!')


def ask_and_execute_vacancies():
    """
    Получает пользовательские данные и испольняет код для вывода таблицы с вакансиями
    :return: None
    """
    file_name = input('Введите название файла: ')
    filter_string = input('Введите параметр фильтрации: ')
    sort_column = input('Введите параметр сортировки: ')
    sort_reversed_string = input('Обратный порядок сортировки (Да / Нет): ')
    number_string = input('Введите диапазон вывода: ')
    columns_string = input('Введите требуемые столбцы: ')

    execute_vacancies(file_name, filter_string, sort_column, sort_reversed_string, number_string, columns_string)


def ask_and_execute_reports():
    """
    Получает пользовательские данные и испольняет код для получение рапорта в различных видах
    :return: None
    """
    file_name = input('Введите название файла: ')
    prof_name = input('Введите название профессии: ')
    execute_reports(file_name, prof_name)


def execute_program():
    """
    Запускает парсер
    :return: None
    """
    ask_for = input('Что вам нужно? (Вакансии/Статистика): ')
    if ask_for.lower() == 'вакансии':
        ask_and_execute_vacancies()
    elif ask_for.lower() == 'статистика':
        ask_and_execute_reports()
    else:
        print('Такого мы пока еще не разработали. Перезапустите программу и введите одно из возможных значений.')


if __name__ == "__main__":
    execute_program()
